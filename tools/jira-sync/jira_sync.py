#!/usr/bin/env python3
"""
Jira Ticket Sync Script
=======================
Fetches all Jira tickets ever assigned to you, organizes them by status into
folders (ToDo / InProgress / InReview), and writes a rich summary for each ticket.

Each run is idempotent:
  - Updated tickets are overwritten in-place (even if they moved folders)
  - Tickets no longer returned by the query are removed (stale cleanup)

Run:
    python jira_sync.py

Requirements:
    pip install -r requirements.txt
"""

import calendar
import json
import math
import os
import re
import shutil
import subprocess
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

import requests
import pathlib
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# Load .env from beside this script first, so the tool works from any working directory
# (it is vendored into VivaAerobus.Generic.ApiLLM/tools/jira-sync). Fall back to the
# default upward search for the original standalone layout.
load_dotenv(pathlib.Path(__file__).with_name(".env"))
load_dotenv()

JIRA_BASE_URL: str = (os.getenv("JIRA_BASE_URL") or "").rstrip("/")
JIRA_EMAIL: str = os.getenv("JIRA_EMAIL") or ""
JIRA_API_TOKEN: str = os.getenv("JIRA_API_TOKEN") or ""
JIRA_PROJECT_KEY: str = os.getenv("JIRA_PROJECT_KEY") or ""   # optional
OUTPUT_DIR: Path = Path(
    os.getenv("OUTPUT_DIR") or Path(__file__).parent / "jira_tickets"
)

# GitHub repo used for PR lookups  (e.g. "MyOrg/MyRepo")
GITHUB_REPO: str = os.getenv("GITHUB_REPO") or ""

# When True the script will automatically:
#   - transition a ticket to UAT when its PR is merged
#   - unassign the ticket from the current user
#   - add a "Daniel" label so tickets stay trackable after reassignment
AUTO_TRANSITION_UAT: bool = os.getenv("AUTO_TRANSITION_UAT", "false").lower() == "true"

# Label appended to every ticket to keep them filterable after unassigning.
# OPT-IN: blank (the default) disables this Jira write entirely — the plain
# sync verb must stay read-only unless the user explicitly configures a label.
TRACKING_LABEL: str = os.getenv("TRACKING_LABEL", "")

# ---------------------------------------------------------------------------
# Status → folder mapping  (lower-case keys)
# ---------------------------------------------------------------------------

STATUS_FOLDER_MAP: dict[str, str] = {
    # ToDo
    "to do": "ToDo",
    "open": "ToDo",
    "backlog": "ToDo",
    "selected for development": "ToDo",
    "ready for sprint": "ToDo",
    "new": "ToDo",
    # InProgress
    "in progress": "InProgress",
    "in development": "InProgress",
    "development": "InProgress",
    "doing": "InProgress",
    # InReview
    "in review": "InReview",
    "code review": "InReview",
    "ready for review": "InReview",
    "peer review": "InReview",
    "review": "InReview",
    "feedback": "InReview",
    "in feedback": "InReview",
    "changes requested": "InReview",
    "needs work": "InReview",
    # UAT
    "uat": "UAT",
    "user acceptance": "UAT",
    "user acceptance testing": "UAT",
    "testing": "UAT",
    "qa": "UAT",
    "qa review": "UAT",
    "ready for uat": "UAT",
    # Released / Done
    "done": "Released",
    "closed": "Released",
    "resolved": "Released",
    "released": "Released",
    "merged": "Released",
    "completed": "Released",
    "accepted": "Released",
    "in production": "Released",
    "production": "Released",
    "deployed": "Released",
    "ready for prod": "Released",
    "ready for production": "Released",
    "ready to deploy": "Released",
    "ready for deployment": "Released",
}

ALL_FOLDERS = ["ToDo", "InProgress", "InReview", "UAT", "Released", "Others"]
ALLOWED_ATTACHMENT_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".txt", ".md"}

DIVIDER = "-" * 70


# ===========================================================================
# Jira API helpers
# ===========================================================================

def _auth() -> tuple[str, str]:
    return (JIRA_EMAIL, JIRA_API_TOKEN)


def _headers() -> dict:
    return {"Accept": "application/json", "Content-Type": "application/json"}


def _get(path: str, params: dict | None = None) -> dict:
    url = f"{JIRA_BASE_URL}{path}"
    resp = requests.get(url, auth=_auth(), headers=_headers(), params=params or {}, timeout=30)
    resp.raise_for_status()
    return resp.json()


def _post(path: str, body: dict) -> dict:
    url = f"{JIRA_BASE_URL}{path}"
    resp = requests.post(url, auth=_auth(), headers=_headers(), json=body, timeout=30)
    resp.raise_for_status()
    return resp.json()


def fetch_changelog(issue_key: str) -> list[dict]:
    """Fetch the full changelog for a single issue via the individual issue endpoint."""
    histories: list[dict] = []
    start_at = 0
    while True:
        data = _get(
            f"/rest/api/3/issue/{issue_key}/changelog",
            params={"startAt": start_at, "maxResults": 100},
        )
        batch = data.get("values", [])
        histories.extend(batch)
        total = data.get("total", 0)
        start_at += len(batch)
        if start_at >= total or not batch:
            break
    return histories


def fetch_single_issue(issue_key: str) -> dict:
    """Fetch any single Jira issue by key with full fields + changelog."""
    data = _get(
        f"/rest/api/3/issue/{issue_key}",
        params={
            "fields": "summary,description,status,assignee,created,"
                      "timeoriginalestimate,timespent,worklog,comment,attachment,labels,"
                      "subtasks",
            "expand": "changelog",
        },
    )
    issue: dict = {"key": data["key"], "fields": data["fields"]}
    # expand=changelog embeds histories; fall back to dedicated endpoint if empty
    changelog_histories = data.get("changelog", {}).get("histories", [])
    if not changelog_histories:
        try:
            changelog_histories = fetch_changelog(issue_key)
        except Exception:
            changelog_histories = []
    issue["changelog"] = {"histories": changelog_histories}
    return issue


def fetch_subtask_details(subtasks: list[dict]) -> list[dict]:
    """
    Fetch summary, status and DESCRIPTION for each subtask of an issue.

    The parent issue embeds only key/summary/status for its subtasks, but under
    the current workflow the test-case matrix and the evidence checklist live in
    the subtask *description* (e.g. API-1779 on API-1584). Without this the
    acceptance criteria are invisible to the sync, which is the whole point of
    pulling the ticket down.
    """
    out: list[dict] = []
    for st in subtasks or []:
        key = st.get("key")
        if not key:
            continue
        st_fields = st.get("fields") or {}
        entry = {
            "key": key,
            "summary": st_fields.get("summary", ""),
            "status": (st_fields.get("status") or {}).get("name", "Unknown"),
            "description": "",
        }
        try:
            data = _get(
                f"/rest/api/3/issue/{key}",
                params={"fields": "summary,status,description"},
            )
            f = data.get("fields", {})
            entry["summary"] = f.get("summary") or entry["summary"]
            entry["status"] = (f.get("status") or {}).get("name") or entry["status"]
            entry["description"] = field_to_text(f.get("description"))
        except Exception as exc:  # never let one subtask break the whole sync
            entry["description"] = f"(could not fetch subtask: {exc})"
        out.append(entry)
    return out


def fetch_all_comments(issue_key: str, comment_field: dict | None) -> list[dict]:
    """
    Return every comment for an issue, paginating when necessary.

    The `comment` field embedded in a search/issue response is capped by Jira: it
    reports `total` next to a possibly shorter `comments` list. Anything past the
    cap is the *newest* activity — exactly where test cases and sign-off land —
    so relying on the embedded list alone drops them with no warning.
    """
    field = comment_field or {}
    embedded = field.get("comments") or []
    total = field.get("total")
    if total is None or int(total) <= len(embedded):
        return embedded

    collected: list[dict] = []
    start_at = 0
    while True:
        data = _get(
            f"/rest/api/3/issue/{issue_key}/comment",
            params={"startAt": start_at, "maxResults": 100, "orderBy": "created"},
        )
        batch = data.get("comments") or []
        collected.extend(batch)
        start_at += len(batch)
        if not batch or start_at >= int(data.get("total", 0)):
            break
    return collected or embedded


def fetch_all_assigned_issues(
    date_from: str | None = None,
    date_to: str | None = None,
) -> list[dict]:
    """Return issues ever assigned to the current user, optionally filtered by update date range.

    Args:
        date_from: ISO date string (YYYY-MM-DD) – only tickets updated on or after this date.
        date_to:   ISO date string (YYYY-MM-DD) – only tickets updated on or before this date.
    """
    issues: list[dict] = []
    next_page_token: str | None = None
    max_results = 50

    project_filter = f" AND project = {JIRA_PROJECT_KEY}" if JIRA_PROJECT_KEY else ""
    date_filter = ""
    if date_from:
        date_filter += f' AND updated >= "{date_from}"'
    if date_to:
        date_filter += f' AND updated <= "{date_to}"'
    jql = f"assignee was currentUser(){project_filter}{date_filter} ORDER BY updated DESC"

    page = 0
    while True:
        body: dict = {
            "jql": jql,
            "maxResults": max_results,
            "fields": [
                "summary", "description", "status", "assignee", "created",
                "timeoriginalestimate", "timespent", "worklog", "comment", "attachment",
                "labels", "subtasks",
            ],
        }
        if next_page_token:
            body["nextPageToken"] = next_page_token

        data = _post("/rest/api/3/search/jql", body=body)
        batch = data.get("issues", [])
        issues.extend(batch)
        page += 1
        is_last = data.get("isLast", True)
        next_page_token = data.get("nextPageToken")
        print(f"  Fetched {len(issues)} tickets (page {page})...", end="\r", flush=True)
        if is_last or not batch or not next_page_token:
            break

    print()  # newline after progress

    # Attach changelog to each issue (fetched separately via dedicated endpoint)
    print("  Fetching changelogs...", end="\r", flush=True)
    for i, issue in enumerate(issues):
        try:
            histories = fetch_changelog(issue["key"])
            issue.setdefault("changelog", {})["histories"] = histories
        except Exception:
            issue.setdefault("changelog", {})["histories"] = []
        print(f"  Changelogs: {i + 1}/{len(issues)}    ", end="\r", flush=True)
    print()

    return issues


def fetch_all_worklogs(issue_key: str) -> list[dict]:
    """Fetch complete worklog list for an issue (handles pagination)."""
    logs: list[dict] = []
    start_at = 0
    while True:
        data = _get(f"/rest/api/3/issue/{issue_key}/worklog", params={"startAt": start_at})
        batch = data.get("worklogs", [])
        logs.extend(batch)
        total = data.get("total", 0)
        start_at += len(batch)
        if start_at >= total or not batch:
            break
    return logs


# ---------------------------------------------------------------------------
# Worklog write helpers
# ---------------------------------------------------------------------------

def download_attachment(url: str, dest: Path) -> None:
    resp = requests.get(url, auth=_auth(), timeout=60, stream=True)
    resp.raise_for_status()
    with open(dest, "wb") as fh:
        for chunk in resp.iter_content(chunk_size=16_384):
            fh.write(chunk)


# ===========================================================================
# Text / ADF helpers
# ===========================================================================

# Node types that only need a block break appended after their children.
# Types with explicit handlers below (heading, lists, tables, codeBlock, panel…)
# must NOT be listed here or they would emit a duplicate separator.
_ADF_BLOCK_TYPES = ("paragraph", "blockquote", "listItem")


def _adf_cell_text(node) -> str:
    """Render one table cell onto a single line (internal breaks become ' / ')."""
    return re.sub(r"\s*\n+\s*", " / ", adf_to_text(node)).strip()


def _adf_table_to_text(node) -> str:
    """
    Render an ADF table.

    Emitted as one labelled block per row rather than pipe-joined columns: the
    test-case matrices live in tables that are 8-9 columns wide, and a pipe join
    of those is unreadable. Losing which value belongs to which column is the
    exact failure this replaces — previously every cell landed on its own line
    with no row or column marker, so a matrix was indistinguishable from prose.
    """
    rows: list[tuple[list[str], bool]] = []
    for row in node.get("content", []) or []:
        if row.get("type") != "tableRow":
            continue
        cells: list[str] = []
        is_header = False
        for cell in row.get("content", []) or []:
            if cell.get("type") == "tableHeader":
                is_header = True
            cells.append(_adf_cell_text(cell))
        if any(cells):
            rows.append((cells, is_header))

    if not rows:
        return ""

    headers: list[str] = []
    body = rows
    # Jira frequently leaves the first row as plain tableCell rather than
    # tableHeader, so fall back to treating row 1 as the header whenever the
    # table is wide enough that a pipe join would be unreadable.
    if rows[0][1] or (len(rows) > 1 and len(rows[0][0]) >= 4):
        headers = rows[0][0]
        body = rows[1:]

    out: list[str] = []
    if headers and body:
        width = max((len(h) for h in headers if h), default=0)
        for idx, (cells, _) in enumerate(body, start=1):
            out.append(f"  Row {idx}:")
            for col, value in enumerate(cells):
                label = (headers[col] if col < len(headers) else "") or f"col{col + 1}"
                out.append(f"    {label:<{width}} : {value}")
            out.append("")
    else:
        for cells, is_header in rows:
            out.append("  | " + " | ".join(cells) + " |")
            if is_header:
                out.append("  |" + "|".join("-" * (len(c) + 2) for c in cells) + "|")
    return "\n".join(out)


def adf_to_text(node, _buf: list | None = None) -> str:
    """Recursively convert Atlassian Document Format (ADF) to plain text."""
    if _buf is None:
        _buf = []

    if isinstance(node, dict):
        ntype = node.get("type", "")
        attrs = node.get("attrs", {}) or {}
        if ntype == "text":
            _buf.append(node.get("text", ""))
        elif ntype in ("hardBreak", "rule"):
            _buf.append("\n")
        elif ntype == "mention":
            # attrs.text usually already carries the leading '@' — don't double it.
            _buf.append("@" + str(attrs.get("text") or "someone").lstrip("@"))
        elif ntype == "emoji":
            _buf.append(attrs.get("shortName", ""))
        elif ntype == "status":
            _buf.append(f"[{attrs.get('text', '')}]")
        elif ntype == "date":
            _buf.append(str(attrs.get("timestamp", "")))
        elif ntype == "inlineCard":
            _buf.append(attrs.get("url", ""))
        elif ntype in ("blockCard", "embedCard"):
            _buf.append(f"{attrs.get('url', '')}\n")
        elif ntype in ("media", "mediaInline"):
            # Previously dropped entirely, silently losing screenshots that the
            # test cases refer to ("the methods shown in the TC01 image").
            label = attrs.get("alt") or attrs.get("id") or "attachment"
            _buf.append(f"[image: {label}]\n")
        elif ntype == "heading":
            level = int(attrs.get("level", 1) or 1)
            _buf.append("\n" + "#" * level + " ")
            for child in node.get("content", []) or []:
                adf_to_text(child, _buf)
            _buf.append("\n")
        elif ntype in ("bulletList", "orderedList"):
            ordered = ntype == "orderedList"
            start = int(attrs.get("order", 1) or 1) if ordered else 1
            for i, item in enumerate(node.get("content", []) or []):
                inner = adf_to_text(item).strip()
                if not inner:
                    continue
                marker = f"{start + i}. " if ordered else "- "
                item_lines = inner.split("\n")
                _buf.append(marker + item_lines[0] + "\n")
                for extra in item_lines[1:]:
                    if extra.strip():
                        _buf.append("  " + extra + "\n")
            _buf.append("\n")
        elif ntype == "codeBlock":
            tmp: list = []
            for child in node.get("content", []) or []:
                adf_to_text(child, tmp)
            lang = attrs.get("language") or ""
            _buf.append(f"\n```{lang}\n" + "".join(tmp).strip() + "\n```\n")
        elif ntype == "table":
            _buf.append("\n" + _adf_table_to_text(node) + "\n")
        elif ntype in ("panel", "expand", "nestedExpand"):
            # expand/panel titles carry the label that identifies a block as the
            # test cases; dropping attrs.title made the block anonymous.
            title = attrs.get("title") or attrs.get("panelType") or ntype
            _buf.append(f"\n[{title}]\n")
            for child in node.get("content", []) or []:
                adf_to_text(child, _buf)
            _buf.append("\n")
        else:
            for child in node.get("content", []) or []:
                adf_to_text(child, _buf)
            if ntype in _ADF_BLOCK_TYPES:
                _buf.append("\n")
    elif isinstance(node, list):
        for item in node:
            adf_to_text(item, _buf)

    return re.sub(r"\n{3,}", "\n\n", "".join(_buf)).strip()


def strip_markdown(text: str) -> str:
    """Best-effort markdown → plain-text conversion."""
    if not text:
        return ""
    text = re.sub(r"```[\s\S]*?```", lambda m: re.sub(r"```", "", m.group()), text)
    text = re.sub(r"`([^`]+)`", r"\1", text)
    text = re.sub(r"^#{1,6}\s+", "", text, flags=re.MULTILINE)
    text = re.sub(r"\*{1,3}([^*\n]+)\*{1,3}", r"\1", text)
    text = re.sub(r"_{1,3}([^_\n]+)_{1,3}", r"\1", text)
    text = re.sub(r"!\[([^\]]*)\]\([^\)]+\)", r"[image: \1]", text)
    text = re.sub(r"\[([^\]]+)\]\([^\)]+\)", r"\1", text)
    text = re.sub(r"^>\s*", "", text, flags=re.MULTILINE)
    text = re.sub(r"^[-*_]{3,}\s*$", "", text, flags=re.MULTILINE)
    text = re.sub(r"^\s*[-*+]\s+", "• ", text, flags=re.MULTILINE)
    text = re.sub(r"^\s*\d+\.\s+", "", text, flags=re.MULTILINE)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def field_to_text(value) -> str:
    """Convert a Jira field value (ADF dict, plain string, or None) to text."""
    if value is None:
        return ""
    if isinstance(value, dict):
        return adf_to_text(value)
    return strip_markdown(str(value))


# ===========================================================================
# Domain helpers
# ===========================================================================

def folder_for_status(status: str) -> str:
    return STATUS_FOLDER_MAP.get(status.lower(), "ToDo")


# Labels that drive the test-case workflow. These are the real gate — a ticket's
# Jira *status* says nothing about whether its test cases exist or have been run.
TEST_CASE_READY_LABEL = "TestCaseReady"
TEST_COMPLETE_LABEL = "TestComplete"

# Every ticket in this project carries exactly one subtask with this summary; it holds
# the test-case matrix and the evidence attachments.
EVIDENCE_SUBTASK_SUMMARY = "Test Cases, execution and evidences"

# Status the parent moves to once the PR is open and evidence is filed.
IN_REVIEW_STATUS = "In review"


def describe_test_flow(labels: list[str]) -> str:
    """
    Summarise where a ticket sits in the test-case workflow.

    TestCaseReady  = a comment/subtask carries the test cases as acceptance criteria.
    TestComplete   = every test case was run and evidence was filed on the subtask.
    """
    lower = {str(lbl).lower() for lbl in labels or []}
    ready = TEST_CASE_READY_LABEL.lower() in lower
    complete = TEST_COMPLETE_LABEL.lower() in lower
    if complete and ready:
        return "TEST COMPLETE (test cases defined, all run + evidence filed)"
    if complete:
        return f"TEST COMPLETE (warning: no {TEST_CASE_READY_LABEL} label)"
    if ready:
        return "TEST CASES READY — execution + evidence still pending"
    return f"not started (neither {TEST_CASE_READY_LABEL} nor {TEST_COMPLETE_LABEL})"


def seconds_to_human(seconds: int | float | None) -> str:
    if not seconds:
        return "Not set"
    h = seconds / 3600
    if h < 8:
        return f"{h:.1f}h"
    d = h / 8
    return f"{d:.1f}d  ({h:.0f}h)"


def parse_jira_date(date_str: Optional[str]) -> Optional[datetime]:
    if not date_str:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str[:26], fmt[:len(fmt)])
        except ValueError:
            continue
    return None


def get_start_date(issue: dict) -> Optional[datetime]:
    """Return the datetime when the issue was first moved to an In-Progress state."""
    histories = sorted(
        issue.get("changelog", {}).get("histories", []),
        key=lambda h: h.get("created", ""),
    )
    for history in histories:
        for item in history.get("items", []):
            if item.get("field") == "status":
                if "progress" in (item.get("toString") or "").lower():
                    return parse_jira_date(history.get("created"))
    # Fall back to issue creation date
    return parse_jira_date(issue.get("fields", {}).get("created"))


def count_devoluciones(issue: dict) -> int:
    """
    Count 'devoluciones': transitions where a ticket moves from InReview
    back to InProgress (i.e., it was rejected and needs more work).
    """
    count = 0
    histories = issue.get("changelog", {}).get("histories", [])
    for history in histories:
        for item in history.get("items", []):
            if item.get("field") == "status":
                from_s = (item.get("fromString") or "").lower()
                to_s = (item.get("toString") or "").lower()
                if "review" in from_s and "progress" in to_s:
                    count += 1
    return count


# ---------------------------------------------------------------------------
# Completion-like statuses (beyond the three tracked folders)
# ---------------------------------------------------------------------------
_DONE_KEYWORDS = ("released", "done", "closed", "resolved", "merged", "completed", "accepted", "in production", "production", "deployed", "ready for prod", "ready for production", "ready to deploy", "ready for deployment")
_UAT_KEYWORDS  = ("uat", "user acceptance", "testing", "qa")


def _is_done_status(status_name: str) -> bool:
    sl = status_name.lower()
    return any(k in sl for k in _DONE_KEYWORDS)


def _is_uat_status(status_name: str) -> bool:
    sl = status_name.lower()
    return any(k in sl for k in _UAT_KEYWORDS)


def _ever_reached_review(issue: dict) -> bool:
    """True if the issue has ever been in a review-like state."""
    for history in issue.get("changelog", {}).get("histories", []):
        for item in history.get("items", []):
            if item.get("field") == "status":
                if "review" in (item.get("toString") or "").lower():
                    return True
    return False


def compute_avance(estimated_seconds: int, total_logged: int, status_name: str) -> tuple[int, str]:
    """
    Returns (avance_pct: int, desviacion: str).
    avance_pct is capped at 100 for display; deviation is shown separately.
    """
    if _is_done_status(status_name) or _is_uat_status(status_name):
        return 100, "NA"

    if estimated_seconds <= 0:
        # No estimate — if any hours logged call it 100%, else 0%
        pct = 100 if total_logged > 0 else 0
        return pct, "NA"

    raw_pct = round(total_logged / estimated_seconds * 100)
    avance = min(raw_pct, 100)

    if raw_pct > 100:
        over = raw_pct - 100
        desviacion = f"+{over}%"
    else:
        desviacion = "NA"

    return avance, desviacion


def determine_proximo_paso(
    folder: str,
    status_name: str,
    avance_pct: int,
    devoluciones: int,
    ever_in_review: bool,
    pr_approved: bool = False,
    pr_merged: bool = False,
) -> str:
    """
    Priority order:
    1. Released / prod         → NA
    2. UAT                     → UAT messages
    3. InReview
       a. PR merged (but still in review) → being moved to UAT
       b. PR approved + open              → approved, waiting to merge
       c. devoluciones > 0               → waiting after fixes
       d. first pass                     → waiting for review
    4. InProgress
       a. devoluciones > 0  → applying reviewer feedback
       b. 0%  → starting
       c. <40% → initiating
       d. <75% → continuing
       e. <100% → finishing
       f. 100%+ → finalizing
    5. ToDo
       a. regression → revisiting
       b. not started → pending
    """
    if _is_done_status(status_name):
        return "NA"

    if folder == "UAT" or _is_uat_status(status_name):
        return "Ticket en proceso de validación UAT"

    if folder == "InReview":
        if pr_merged:
            return "Ticket mezclado, moviendo a UAT"
        if pr_approved:
            return "Ticket revisado y aprobado, esperando ser mezclado"
        if devoluciones > 0:
            return "Esperando aprobación final tras correcciones realizadas"
        return "En espera de revisión por parte del equipo"

    if folder == "InProgress":
        if devoluciones > 0:
            return "Haciendo correcciones a comentarios hechos por el revisor"
        if avance_pct == 0:
            return "Revisando la lógica y el código para iniciar el desarrollo"
        if avance_pct < 40:
            return "Iniciando el desarrollo"
        if avance_pct < 75:
            return "Continuar el desarrollo e iniciar pruebas"
        if avance_pct < 100:
            return "Finalizando implementación y preparando para revisión"
        return "Completando detalles finales y enviando a revisión"

    # ToDo
    if ever_in_review:
        return "Revisando elementos antes de reiniciar el desarrollo"
    if avance_pct == 0:
        return "Pendiente de inicio — ticket en cola de desarrollo"
    return "Revisando la lógica y el código para iniciar el desarrollo"


def build_summary_entry(
    issue: dict,
    folder: str,
    avance_pct: int,
    desviacion: str,
    devoluciones: int,
    pr_approved: bool = False,
    pr_merged: bool = False,
) -> str:
    """Build the formatted block for a single ticket in summary.txt."""
    fields = issue.get("fields", {})
    key = issue["key"]
    title = fields.get("summary") or "No title"
    status = fields.get("status", {}).get("name") or "Unknown"

    ever_in_review = _ever_reached_review(issue)
    proximo_paso = determine_proximo_paso(
        folder, status, avance_pct, devoluciones, ever_in_review,
        pr_approved=pr_approved, pr_merged=pr_merged,
    )

    avance_display = f"{avance_pct}%"
    if _is_done_status(status):
        avance_display = "100% (merged)"
    elif _is_uat_status(status):
        avance_display = "100% (en UAT)"

    lines = [
        f"Ticket {key} {title}",
        f"\tAvance: {avance_display}",
        f"\tEstado: {status}",
        f"\tDevoluciones: {devoluciones}",
        f"\tDesviación: {desviacion}",
        f"\tBloqueo/Riesgo: NA",
        f"\tPróximo Paso: {proximo_paso}",
        "",
    ]
    return "\n".join(lines)


def write_summary(output_dir: Path, entries: list[str]) -> None:
    """Write summary.txt to the top-level output folder."""
    summary_path = output_dir / "summary.txt"
    header = [
        "RESUMEN DE TICKETS",
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "=" * 70,
        "",
    ]
    summary_path.write_text("\n".join(header) + "\n" + "\n".join(entries), encoding="utf-8")
    print(f"\nSummary written → {summary_path}")


# ===========================================================================
# GitHub PR helpers
# ===========================================================================

_PR_URL_RE = re.compile(r"github\.com/([\w.-]+/[\w.-]+)/pull/(\d+)", re.IGNORECASE)
_PR_NUM_RE = re.compile(r"Pull Request #(\d+)\s*·\s*([\w.-]+/[\w.-]+)", re.IGNORECASE)


def extract_pr_from_comments(comments: list[dict]) -> tuple[Optional[str], Optional[int]]:
    """
    Scan all comment bodies for a GitHub PR reference.
    Returns (repo, pr_number) or (None, None).
    """
    for c in comments:
        body = field_to_text(c.get("body") or {})
        # Match full GitHub URL
        m = _PR_URL_RE.search(body)
        if m:
            return m.group(1), int(m.group(2))
        # Match "Pull Request #NNN · Org/Repo" pattern (Jira inline cards)
        m = _PR_NUM_RE.search(body)
        if m:
            return m.group(2), int(m.group(1))
    return None, None


def get_pr_state(repo: str, pr_number: int) -> dict:
    """
    Query GitHub via the `gh` CLI.
    Returns a dict with keys: state, review_decision, is_merged, is_approved.
    Falls back to empty state on any error.
    """
    empty = {"state": "", "review_decision": "", "is_merged": False, "is_approved": False}
    if not repo or not pr_number:
        return empty
    try:
        result = subprocess.run(
            ["gh", "pr", "view", str(pr_number), "--repo", repo,
             "--json", "state,reviewDecision,mergedAt"],
            capture_output=True, text=True, timeout=20,
        )
        if result.returncode != 0:
            return empty
        data = json.loads(result.stdout)
        state = (data.get("state") or "").upper()          # OPEN / CLOSED / MERGED
        review_decision = (data.get("reviewDecision") or "").upper()  # APPROVED / CHANGES_REQUESTED / ...
        is_merged = bool(data.get("mergedAt"))
        is_approved = review_decision == "APPROVED"
        return {"state": state, "review_decision": review_decision,
                "is_merged": is_merged, "is_approved": is_approved}
    except Exception:
        return empty


# ===========================================================================
# Jira write helpers  (only invoked when AUTO_TRANSITION_UAT=true)
# ===========================================================================

def _jira_put(path: str, body: dict) -> None:
    url = f"{JIRA_BASE_URL}{path}"
    resp = requests.put(url, auth=_auth(), headers=_headers(), json=body, timeout=20)
    resp.raise_for_status()


def _find_uat_transition_id(issue_key: str) -> Optional[str]:
    """Return the Jira transition ID whose name contains 'UAT' or 'Testing'."""
    data = _get(f"/rest/api/3/issue/{issue_key}/transitions")
    for t in data.get("transitions", []):
        name = (t.get("name") or "").lower()
        if "uat" in name or "testing" in name or "acceptance" in name:
            return t["id"]
    return None


def jira_transition_to_uat(issue_key: str) -> bool:
    """Transition a ticket to UAT. Returns True on success."""
    tid = _find_uat_transition_id(issue_key)
    if not tid:
        print(f"    ⚠  No UAT transition found for {issue_key}")
        return False
    try:
        url = f"{JIRA_BASE_URL}/rest/api/3/issue/{issue_key}/transitions"
        resp = requests.post(url, auth=_auth(), headers=_headers(),
                             json={"transition": {"id": tid}}, timeout=20)
        resp.raise_for_status()
        return True
    except Exception as exc:
        print(f"    ⚠  Transition to UAT failed for {issue_key}: {exc}")
        return False


def jira_unassign(issue_key: str) -> None:
    """Remove the assignee from a Jira ticket."""
    try:
        _jira_put(f"/rest/api/3/issue/{issue_key}", {"fields": {"assignee": None}})
    except Exception as exc:
        print(f"    ⚠  Unassign failed for {issue_key}: {exc}")


def jira_ensure_label(issue_key: str, current_labels: list[str], label: str) -> None:
    """Add a label to a ticket if it is not already present."""
    if label in current_labels:
        return
    try:
        new_labels = current_labels + [label]
        _jira_put(f"/rest/api/3/issue/{issue_key}", {"fields": {"labels": new_labels}})
    except Exception as exc:
        print(f"    ⚠  Label update failed for {issue_key}: {exc}")


# ===========================================================================
# Ticket delivery — the write side of the dev-ticket lifecycle
# (see VivaAerobus.Generic.ApiLLM/llm/TICKET-DELIVERY.md)
# ===========================================================================

def jira_update_labels(issue_key: str, add: list[str] | None = None,
                       remove: list[str] | None = None) -> list[str]:
    """
    Add and/or remove labels, preserving everything else. Returns the new list.

    Uses the labels `update` verb rather than a `fields` overwrite so a concurrent
    edit cannot be clobbered by a stale read.
    """
    ops: list[dict] = []
    for lbl in add or []:
        ops.append({"add": lbl})
    for lbl in remove or []:
        ops.append({"remove": lbl})
    if not ops:
        return []
    _jira_put(f"/rest/api/3/issue/{issue_key}", {"update": {"labels": ops}})
    data = _get(f"/rest/api/3/issue/{issue_key}", params={"fields": "labels"})
    return data.get("fields", {}).get("labels", [])


def jira_list_transitions(issue_key: str) -> list[tuple[str, str]]:
    """Return [(id, name)] of the transitions currently available on the issue."""
    data = _get(f"/rest/api/3/issue/{issue_key}/transitions")
    return [(t["id"], t.get("name", "")) for t in data.get("transitions", [])]


def jira_transition(issue_key: str, target: str) -> bool:
    """
    Transition an issue to the named target status (case-insensitive, substring).

    Jira exposes only the transitions valid from the CURRENT status, so a failure
    here usually means the workflow does not allow that jump from where the issue
    is — not that the name is wrong. The available names are printed to say which.
    """
    available = jira_list_transitions(issue_key)
    wanted = target.strip().lower()
    tid = next((i for i, n in available if n.strip().lower() == wanted), None)
    if tid is None:
        tid = next((i for i, n in available if wanted in n.strip().lower()), None)
    if tid is None:
        names = ", ".join(n for _, n in available) or "(none)"
        print(f"    ⚠  No transition matching '{target}' on {issue_key}. Available: {names}")
        return False
    try:
        resp = requests.post(
            f"{JIRA_BASE_URL}/rest/api/3/issue/{issue_key}/transitions",
            auth=_auth(), headers=_headers(), json={"transition": {"id": tid}}, timeout=20)
        resp.raise_for_status()
        return True
    except Exception as exc:
        print(f"    ⚠  Transition '{target}' failed for {issue_key}: {exc}")
        return False


def jira_add_comment(issue_key: str, text: str) -> bool:
    """
    Post a plain-text comment. Jira Cloud v3 requires ADF, so the text is wrapped;
    blank lines separate paragraphs.
    """
    paragraphs = [p for p in text.split("\n\n")]
    body = {
        "body": {
            "type": "doc", "version": 1,
            "content": [
                {"type": "paragraph",
                 "content": [{"type": "text", "text": p}] if p.strip() else []}
                for p in paragraphs
            ],
        }
    }
    try:
        _post(f"/rest/api/3/issue/{issue_key}/comment", body=body)
        return True
    except Exception as exc:
        print(f"    ⚠  Comment failed for {issue_key}: {exc}")
        return False


def jira_upload_attachments(issue_key: str, paths: list[Path]) -> list[str]:
    """
    Upload files to an issue. Returns the filenames Jira accepted.

    Note the two non-obvious requirements of this endpoint: the
    `X-Atlassian-Token: no-check` header (XSRF opt-out) and NO Content-Type header
    — requests must set the multipart boundary itself.
    """
    url = f"{JIRA_BASE_URL}/rest/api/3/issue/{issue_key}/attachments"
    headers = {"Accept": "application/json", "X-Atlassian-Token": "no-check"}
    uploaded: list[str] = []
    for path in paths:
        if not path.is_file():
            print(f"    ⚠  Not a file, skipped: {path}")
            continue
        try:
            with open(path, "rb") as fh:
                resp = requests.post(url, auth=_auth(), headers=headers,
                                     files={"file": (path.name, fh)}, timeout=120)
            resp.raise_for_status()
            for item in resp.json():
                uploaded.append(item.get("filename", path.name))
            print(f"    ✔ uploaded {path.name}")
        except Exception as exc:
            print(f"    ⚠  Upload failed for {path.name}: {exc}")
    return uploaded


def jira_find_or_create_evidence_subtask(parent_key: str,
                                         summary: str = EVIDENCE_SUBTASK_SUMMARY) -> Optional[str]:
    """
    Return the key of the parent's evidence subtask, creating it when absent.

    Matching is by summary so a subtask created by hand in the UI is reused rather
    than duplicated — every ticket in this project has exactly one.
    """
    data = _get(f"/rest/api/3/issue/{parent_key}", params={"fields": "subtasks,project"})
    fields = data.get("fields", {})
    for sub in fields.get("subtasks") or []:
        if (sub["fields"].get("summary") or "").strip().lower() == summary.strip().lower():
            return sub["key"]

    project_key = (fields.get("project") or {}).get("key")
    if not project_key:
        print(f"    ⚠  Cannot resolve project for {parent_key}; not creating a subtask.")
        return None
    try:
        created = _post("/rest/api/3/issue", body={
            "fields": {
                "project": {"key": project_key},
                "parent": {"key": parent_key},
                "summary": summary,
                "issuetype": {"name": "Subtask"},
            }
        })
        key = created.get("key")
        print(f"    ✔ created evidence subtask {key}")
        return key
    except Exception as exc:
        print(f"    ⚠  Could not create the evidence subtask on {parent_key}: {exc}")
        print("       Create it manually and re-run; everything else still applies.")
        return None


# ===========================================================================
# Domain helpers (continued)
# ===========================================================================

def process_issue(issue: dict, output_dir: Path, folder_override: str | None = None) -> tuple[str, str]:
    """Write the ticket summary file + attachments. Returns (folder_name, summary_entry)."""
    # folder_override forces the ticket into a specific folder (e.g. "Others").
    key = issue["key"]
    fields = issue.get("fields", {})

    # ── Basic fields ────────────────────────────────────────────────────────
    title = fields.get("summary") or "No title"
    status = fields.get("status", {}).get("name") or "Unknown"
    folder_name = folder_override if folder_override else folder_for_status(status)

    assignee_obj = fields.get("assignee")
    current_assignee = (
        assignee_obj.get("displayName", "Unassigned") if assignee_obj else "Unassigned"
    )

    description_text = field_to_text(fields.get("description")) or "No description."

    # ── Dates & estimates ───────────────────────────────────────────────────
    start_dt = get_start_date(issue)
    start_date_str = start_dt.strftime("%Y-%m-%d") if start_dt else "Unknown"

    estimated_seconds: int = fields.get("timeoriginalestimate") or 0
    estimated_display = seconds_to_human(estimated_seconds)

    # Deadline = start + estimated * 1.2  (in calendar days, assuming 8h/day)
    deadline_str = "N/A"
    available_time_str = "N/A"
    if start_dt and estimated_seconds:
        work_hours_with_buffer = estimated_seconds / 3600 * 1.2
        work_days = math.ceil(work_hours_with_buffer / 8)
        deadline_dt = start_dt.replace(tzinfo=None) + timedelta(days=work_days)
        deadline_str = deadline_dt.strftime("%Y-%m-%d")

        now = datetime.now()
        delta = deadline_dt - now
        if delta.total_seconds() >= 0:
            available_time_str = f"{delta.days}d remaining (deadline {deadline_str})"
        else:
            available_time_str = f"OVERDUE by {abs(delta.days)}d (deadline {deadline_str})"

    # ── Worklogs (all team members) ─────────────────────────────────────────
    try:
        worklogs = fetch_all_worklogs(key)
    except Exception:
        # Fall back to inline worklogs from search response
        worklogs = fields.get("worklog", {}).get("worklogs", [])

    time_per_member: dict[str, int] = {}
    for log in worklogs:
        author = log.get("author", {}).get("displayName", "Unknown")
        time_per_member[author] = time_per_member.get(author, 0) + (
            log.get("timeSpentSeconds") or 0
        )
    total_logged = sum(time_per_member.values())

    # ── Devoluciones ────────────────────────────────────────────────────────
    devoluciones = count_devoluciones(issue)

    # ── Avance & Desviación ─────────────────────────────────────────────────
    avance_pct, desviacion = compute_avance(estimated_seconds, total_logged, status)

    # ── PR status via GitHub CLI ─────────────────────────────────────────────
    raw_comments = fetch_all_comments(key, fields.get("comment"))
    pr_repo, pr_number = extract_pr_from_comments(raw_comments)
    effective_repo = pr_repo or GITHUB_REPO   # fall back to env var
    pr_state = get_pr_state(effective_repo, pr_number) if effective_repo and pr_number else {}
    pr_approved = pr_state.get("is_approved", False)
    pr_merged   = pr_state.get("is_merged",   False)

    # ── Auto Jira actions when PR is merged and ticket still in InReview ────
    if AUTO_TRANSITION_UAT and pr_merged and folder_name == "InReview":
        print(f"    → PR #{pr_number} merged, transitioning {key} to UAT...")
        if jira_transition_to_uat(key):
            # Refresh folder/status after transition
            folder_name = "UAT"
            status = "UAT"
            jira_unassign(key)

    # ── Ensure tracking label is present (only when no other labels exist) ───
    if TRACKING_LABEL:
        current_labels = fields.get("labels") or []
        if not current_labels:
            jira_ensure_label(key, current_labels, TRACKING_LABEL)

    # ── Comments ────────────────────────────────────────────────────────────
    comment_blocks: list[str] = []
    for c in raw_comments:
        author = c.get("author", {}).get("displayName", "Unknown")
        created = (c.get("created") or "")[:10]
        body = field_to_text(c.get("body")) or "(empty)"
        comment_blocks.append(f"[{created}] {author}:\n{body}")

    # ── Attachments ─────────────────────────────────────────────────────────
    attachments = [
        a for a in (fields.get("attachment") or [])
        if Path(a.get("filename", "")).suffix.lower() in ALLOWED_ATTACHMENT_EXTS
    ]

    # ── Remove ticket from any other folder it may have lived in previously ──
    for other_folder in ALL_FOLDERS:
        if other_folder == folder_name:
            continue
        old_file = output_dir / other_folder / f"{key}.txt"
        if old_file.exists():
            old_file.unlink()
            print(f"  Moved {key}: {other_folder}/ → {folder_name}/")
        old_att_dir = output_dir / other_folder / f"{key}_attachments"
        if old_att_dir.exists():
            shutil.rmtree(old_att_dir)

    # ── Test-case flow state (driven by labels, not status) ──────────────────
    issue_labels = fields.get("labels") or []
    test_flow_state = describe_test_flow(issue_labels)

    # ── Write output ─────────────────────────────────────────────────────────
    folder_path = output_dir / folder_name
    folder_path.mkdir(parents=True, exist_ok=True)

    lines: list[str] = [
        f"Ticket       : {key}",
        f"Title        : {title}",
        f"Status       : {status}",
        f"Test Flow    : {test_flow_state}",
        f"Labels       : {', '.join(issue_labels) if issue_labels else 'None'}",
        f"Assignee     : {current_assignee}",
        f"Start Date   : {start_date_str}",
        f"Estimated    : {estimated_display}",
        f"Deadline     : {deadline_str}  (start + estimate + 20% buffer)",
        f"Avail. Time  : {available_time_str}",
        f"Devoluciones : {devoluciones}  (InReview→InProgress rejections)",
        f"PR           : #{pr_number} ({effective_repo})" if pr_number else "PR           : None found",
        f"PR Status    : {'MERGED' if pr_merged else 'APPROVED (open)' if pr_approved else pr_state.get('state', 'N/A')}" if pr_state else "PR Status    : N/A",
        "",
        "Time Logged by Team Member:",
        DIVIDER,
    ]
    if time_per_member:
        for member, secs in sorted(time_per_member.items()):
            lines.append(f"  {member:<40} {seconds_to_human(secs)}")
        lines.append(f"  {'TOTAL':<40} {seconds_to_human(total_logged)}")
    else:
        lines.append("  No time logged yet.")

    lines += [
        "",
        "Description:",
        DIVIDER,
        description_text,
        DIVIDER,
        "",
    ]

    subtask_entries = fetch_subtask_details(fields.get("subtasks") or [])
    if subtask_entries:
        lines += ["Subtasks (test cases / evidence):", DIVIDER]
        for st in subtask_entries:
            lines.append(f"  {st['key']}  [{st['status']}]  {st['summary']}")
            if st["description"]:
                for desc_line in st["description"].split("\n"):
                    lines.append(f"    {desc_line}" if desc_line.strip() else "")
            lines.append("")
        lines.append(DIVIDER)
        lines.append("")

    if comment_blocks:
        lines += ["Comments:", DIVIDER]
        for block in comment_blocks:
            lines.append(block)
            lines.append("")
        lines.append(DIVIDER)
    else:
        lines.append("Comments: None")

    if attachments:
        att_dir = folder_path / f"{key}_attachments"
        att_dir.mkdir(exist_ok=True)
        lines += ["", "Attachments:", DIVIDER]
        for idx, att in enumerate(attachments, start=1):
            original_name = att.get("filename", f"file_{idx}")
            ext = Path(original_name).suffix.lower()
            dest_name = f"{key}_{idx}{ext}"
            dest_path = att_dir / dest_name
            try:
                download_attachment(att["content"], dest_path)
                lines.append(f"  [{idx:02d}] {original_name}  →  {dest_name}")
            except Exception as exc:
                lines.append(f"  [{idx:02d}] {original_name}  FAILED: {exc}")
        lines.append(DIVIDER)

    ticket_file = folder_path / f"{key}.txt"
    ticket_file.write_text("\n".join(lines), encoding="utf-8")

    summary_entry = build_summary_entry(
        issue, folder_name, avance_pct, desviacion, devoluciones,
        pr_approved=pr_approved, pr_merged=pr_merged,
    )
    return folder_name, summary_entry


# ===========================================================================
# Stale-ticket cleanup
# ===========================================================================

def cleanup_stale(output_dir: Path, live_keys: set[str]) -> None:
    """Delete ticket files (and their attachment dirs) that are no longer returned."""
    for folder in ALL_FOLDERS:
        folder_path = output_dir / folder
        if not folder_path.exists():
            continue
        for txt_file in folder_path.glob("*.txt"):
            if txt_file.stem not in live_keys:
                txt_file.unlink()
                print(f"  Removed stale: {txt_file.stem}")
                att_dir = folder_path / f"{txt_file.stem}_attachments"
                if att_dir.exists():
                    shutil.rmtree(att_dir)


# ===========================================================================
# Entry point
# ===========================================================================

def _parse_args() -> tuple[str | None, str | None, bool]:
    """Parse CLI arguments for date filtering.

    Usage:
        python jira_sync.py                        # current month (default)
        python jira_sync.py --all                  # all time
        python jira_sync.py --from 2026-01-01      # from date to today
        python jira_sync.py --from 2026-01-01 --to 2026-03-31
    """
    import argparse
    parser = argparse.ArgumentParser(
        description="Sync Jira tickets to local files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--from", dest="date_from", metavar="YYYY-MM-DD",
        help="Only sync tickets updated on or after this date.",
    )
    parser.add_argument(
        "--to", dest="date_to", metavar="YYYY-MM-DD",
        help="Only sync tickets updated on or before this date.",
    )
    parser.add_argument(
        "--all", dest="all_time", action="store_true",
        help="Sync all tickets regardless of date (slow).",
    )
    args = parser.parse_args()

    if args.all_time:
        return None, None, True  # (date_from, date_to, all_time)

    if args.date_from or args.date_to:
        return args.date_from, args.date_to, False

    # Default: current calendar month
    today = date.today()
    date_from = today.strftime("%Y-%m-01")
    date_to = today.strftime("%Y-%m-%d")
    return date_from, date_to, False


def main() -> None:
    # Validate required env vars
    missing = [v for v in ("JIRA_BASE_URL", "JIRA_EMAIL", "JIRA_API_TOKEN") if not os.getenv(v)]
    if missing:
        print("ERROR: Missing required environment variables:", ", ".join(missing))
        print("Fill in your .env file (see .env.example) and re-run.")
        sys.exit(1)

    date_from, date_to, all_time = _parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"Jira base URL : {JIRA_BASE_URL}")
    print(f"User          : {JIRA_EMAIL}")
    if JIRA_PROJECT_KEY:
        print(f"Project filter: {JIRA_PROJECT_KEY}")
    print(f"Output dir    : {OUTPUT_DIR}")
    if all_time:
        print("Date range    : ALL TIME")
    else:
        print(f"Date range    : {date_from or 'beginning'} → {date_to or 'today'}")
    print()

    print("Fetching tickets assigned to you...")
    try:
        issues = fetch_all_assigned_issues(date_from=date_from, date_to=date_to)
    except requests.HTTPError as exc:
        print(f"HTTP error fetching issues: {exc}")
        sys.exit(1)

    print(f"Total tickets : {len(issues)}\n")
    if not issues:
        print("Nothing to do.")
        return

    live_keys: set[str] = set()
    stats: dict[str, int] = {f: 0 for f in ALL_FOLDERS}
    errors: list[str] = []
    summary_entries: list[str] = []   # ordered: most-recently-updated first (JQL order)

    for i, issue in enumerate(issues, start=1):
        key = issue["key"]
        live_keys.add(key)
        try:
            folder, summary_entry = process_issue(issue, OUTPUT_DIR)
            stats[folder] = stats.get(folder, 0) + 1
            summary_entries.append(summary_entry)
            print(f"  [{i:>4}/{len(issues)}] {key:<15} → {folder}/")
        except Exception as exc:
            errors.append(f"{key}: {exc}")
            print(f"  [{i:>4}/{len(issues)}] {key:<15}   ERROR: {exc}")

    if all_time:
        print("\nCleaning up stale tickets...")
        cleanup_stale(OUTPUT_DIR, live_keys)
    else:
        print("\nSkipping stale cleanup (date-filtered sync — only current range was fetched).")

    write_summary(OUTPUT_DIR, summary_entries)

    print("\n" + "=" * 50)
    print("Sync complete.")
    for folder, count in stats.items():
        print(f"  {folder:<14} {count} ticket(s)")
    if errors:
        print(f"\n  Errors ({len(errors)}):")
        for e in errors:
            print(f"    {e}")
    print(f"\nFiles written to: {OUTPUT_DIR}")


# ===========================================================================
# Timesheet / worklog report  (mirrors get-my-jira-info.js logic)
# ===========================================================================

def get_date_range(args: list[str]) -> dict:
    """
    Parse month and year from CLI args (e.g. ``timesheet 4 2026``).
    Defaults to the previous calendar month.
    """
    if len(args) >= 2:
        month = int(args[0])
        year = int(args[1])
    else:
        today = datetime.now()
        first_of_this_month = today.replace(day=1)
        last_month = first_of_this_month - timedelta(days=1)
        month = last_month.month
        year = last_month.year

    last_day = calendar.monthrange(year, month)[1]
    return {
        "month": month,
        "year": year,
        "start_date": f"{year:04d}-{month:02d}-01",
        "end_date": f"{year:04d}-{month:02d}-{last_day:02d}",
    }


def get_current_user() -> dict:
    """Return the authenticated user's Jira profile."""
    return _get("/rest/api/3/myself")


def _search_jql_timesheet(jql: str) -> list[dict]:
    """Run a JQL search for timesheet purposes (lightweight field set, inline worklogs)."""
    issues: list[dict] = []
    next_page_token: str | None = None
    fields = [
        "summary", "status", "assignee", "issuetype",
        "updated", "worklog", "created", "resolutiondate",
        "timeoriginalestimate",
    ]
    while True:
        body: dict = {"jql": jql, "maxResults": 50, "fields": fields}
        if next_page_token:
            body["nextPageToken"] = next_page_token
        data = _post("/rest/api/3/search/jql", body=body)
        batch = data.get("issues", [])
        issues.extend(batch)
        is_last = data.get("isLast", True)
        next_page_token = data.get("nextPageToken")
        if is_last or not batch or not next_page_token:
            break
    return issues


def fetch_issues_with_worklogs_in_range(date_range: dict) -> list[dict]:
    """Issues where the current user logged time within the date range."""
    jql = (
        f'worklogAuthor = currentUser() '
        f'AND worklogDate >= "{date_range["start_date"]}" '
        f'AND worklogDate <= "{date_range["end_date"]}" '
        f'ORDER BY updated DESC'
    )
    return _search_jql_timesheet(jql)


def fetch_assigned_issues_in_range(date_range: dict) -> list[dict]:
    """Issues assigned to the current user at any point during the date range."""
    jql = (
        f'assignee WAS currentUser() '
        f'DURING ("{date_range["start_date"]}", "{date_range["end_date"]}") '
        f'AND (status NOT IN (Done, Canceled) OR resolutiondate >= "{date_range["start_date"]}") '
        f'AND updated >= "{date_range["start_date"]}" '
        f'AND issuetype != Epic '
        f'ORDER BY updated DESC'
    )
    return _search_jql_timesheet(jql)


def calculate_worklog_hours_in_range(
    issue: dict, current_user_id: str, date_range: dict
) -> dict:
    """Count and total hours for worklogs by the current user within the date range."""
    worklogs = (issue.get("fields", {}).get("worklog") or {}).get("worklogs", [])
    filtered = [
        log for log in worklogs
        if (
            log.get("author", {}).get("accountId") == current_user_id
            or log.get("author", {}).get("emailAddress") == current_user_id
            or log.get("author", {}).get("displayName") == current_user_id
        )
        and date_range["start_date"] <= (log.get("started") or "")[:10] <= date_range["end_date"]
    ]
    total_seconds = sum(log.get("timeSpentSeconds", 0) for log in filtered)
    return {"count": len(filtered), "total_hours": f"{total_seconds / 3600:.2f}"}


def organize_worklogs_by_date(
    issues: list[dict], current_user_id: str, date_range: dict
) -> dict:
    """Return ``{ticket_key: {date_str: hours}}`` for the current user within the date range."""
    result: dict = {}
    for issue in issues:
        key = issue["key"]
        worklogs = (issue.get("fields", {}).get("worklog") or {}).get("worklogs", [])
        ticket_data: dict = {}
        for log in worklogs:
            author = log.get("author", {})
            is_me = (
                author.get("accountId") == current_user_id
                or author.get("emailAddress") == current_user_id
                or author.get("displayName") == current_user_id
            )
            if not is_me:
                continue
            log_date = (log.get("started") or "")[:10]
            if date_range["start_date"] <= log_date <= date_range["end_date"]:
                ticket_data[log_date] = (
                    ticket_data.get(log_date, 0) + log.get("timeSpentSeconds", 0) / 3600
                )
        if ticket_data:
            result[key] = ticket_data
    return result


def generate_excel_report(
    issues: list[dict], date_range: dict, current_user_id: str, output_dir: Path
) -> None:
    """Generate a weekly-calendar Excel timesheet matching the JS reference script layout."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    year = date_range["year"]
    month = date_range["month"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}-{year} Timesheet"

    worklog_map = organize_worklogs_by_date(issues, current_user_id, date_range)
    ticket_keys = sorted(worklog_map.keys())

    ticket_info: dict = {}
    for issue in issues:
        ticket_info[issue["key"]] = {
            "summary": issue["fields"].get("summary", ""),
            "status": (issue["fields"].get("status") or {}).get("name", ""),
            "created": issue["fields"].get("created", ""),
        }

    DAY_NAMES = ["SU", "MO", "TU", "WE", "TH", "FR", "SA"]
    gray_fill   = PatternFill(fill_type="solid", fgColor="FFE0E0E0")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFFCC00")
    green_fill  = PatternFill(fill_type="solid", fgColor="FF00FF00")
    blue_fill   = PatternFill(fill_type="solid", fgColor="FFADD8E6")

    month_start = date(year, month, 1)
    month_end   = date(year, month, calendar.monthrange(year, month)[1])

    current_row = 1
    all_total_rows: list[int] = []

    current_date = month_start
    while current_date <= month_end:
        # Find the Sunday that starts this calendar week (JS: getDay() where 0=Sun)
        js_dow = (current_date.weekday() + 1) % 7   # 0=Sun ... 6=Sat
        week_start = current_date - timedelta(days=js_dow)
        week_dates = [week_start + timedelta(days=i) for i in range(7)]

        tickets_this_week = [
            tk for tk in ticket_keys
            if any(worklog_map[tk].get(d.strftime("%Y-%m-%d"), 0) > 0 for d in week_dates)
        ]

        if not tickets_this_week:
            current_date = week_start + timedelta(days=7)
            continue

        # Day-name header row
        ws.cell(current_row, 1).value = ""
        for i, dn in enumerate(DAY_NAMES):
            c = ws.cell(current_row, i + 2)
            c.value = dn
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
        ws.cell(current_row, 9).value = "TOTAL"
        ws.cell(current_row, 9).font = Font(bold=True)
        current_row += 1

        # Separator row
        for i in range(8):
            ws.cell(current_row, i + 2).value = "---"
        current_row += 1

        # Day-number row
        ws.cell(current_row, 1).value = "DAYS"
        ws.cell(current_row, 1).font = Font(bold=True)
        for i, d in enumerate(week_dates):
            c = ws.cell(current_row, i + 2)
            c.value = d.day
            c.alignment = Alignment(horizontal="center")
            if d.month != month:
                c.font = Font(color="FFCCCCCC")
        current_row += 1

        ws.cell(current_row, 1).value = "-----"
        current_row += 1

        ws.cell(current_row, 1).value = "TICKETS"
        ws.cell(current_row, 1).font = Font(bold=True)
        current_row += 1

        ticket_start_row = current_row

        # One row per ticket with hours per day
        for tk in tickets_this_week:
            ws.cell(current_row, 1).value = tk
            for i, d in enumerate(week_dates):
                hours = worklog_map[tk].get(d.strftime("%Y-%m-%d"), 0)
                c = ws.cell(current_row, i + 2)
                if hours > 0:
                    c.value = hours
                    c.number_format = '0.0"h"'
                else:
                    c.value = "-"
                c.alignment = Alignment(horizontal="center")
            total_cell = ws.cell(current_row, 9)
            total_cell.value = f"=SUM(B{current_row}:H{current_row})"
            total_cell.number_format = '0.0"h"'
            total_cell.font = Font(bold=True)
            current_row += 1

        current_row += 1  # blank row

        # Daily total row
        ws.cell(current_row, 1).value = "TOTAL PER DAY"
        ws.cell(current_row, 1).font = Font(bold=True)
        for i in range(7):
            col = get_column_letter(i + 2)
            tc = ws.cell(current_row, i + 2)
            tc.value = f"=SUM({col}{ticket_start_row}:{col}{current_row - 2})"
            tc.number_format = '0.0"h"'
            tc.font = Font(bold=True)
            tc.alignment = Alignment(horizontal="center")
            tc.fill = gray_fill
        week_total = ws.cell(current_row, 9)
        week_total.value = f"=SUM(B{current_row}:H{current_row})"
        week_total.number_format = '0.0"h"'
        week_total.font = Font(bold=True)
        week_total.fill = yellow_fill
        all_total_rows.append(current_row)
        current_row += 2

        current_date = week_start + timedelta(days=7)

    # Monthly grand total
    current_row += 1
    ws.cell(current_row, 1).value = "MONTHLY TOTAL"
    ws.cell(current_row, 1).font = Font(bold=True, size=12)
    grand = ws.cell(current_row, 9)
    grand.value = ("=" + "+".join(f"I{r}" for r in all_total_rows)) if all_total_rows else 0
    grand.number_format = '0.0"h"'
    grand.font = Font(bold=True, size=12)
    grand.fill = green_fill

    # Tickets summary table
    current_row += 3
    ws.cell(current_row, 1).value = "TICKETS ASSIGNED THIS MONTH"
    ws.cell(current_row, 1).font = Font(bold=True, size=12)
    current_row += 2

    for col_idx, header in enumerate(
        ["Ticket", "Title", "", "", "", "", "Date Assigned", "Status", "Total Hours"], start=1
    ):
        c = ws.cell(current_row, col_idx)
        c.value = header
        c.font = Font(bold=True)
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
    current_row += 1

    summary_start = current_row
    sorted_issues = sorted(issues, key=lambda x: x["fields"].get("created", ""))
    RED = "FFFF0000"

    for issue in sorted_issues:
        key = issue["key"]
        info = ticket_info.get(key, {})
        wl = calculate_worklog_hours_in_range(issue, current_user_id, date_range)
        total_h = float(wl["total_hours"])

        ws.cell(current_row, 1).value = key
        ws.cell(current_row, 2).value = info.get("summary", "")
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)

        created_str = info.get("created", "")
        if created_str:
            try:
                created_dt = datetime.fromisoformat(created_str[:19])
                ws.cell(current_row, 7).value = created_dt
                ws.cell(current_row, 7).number_format = "mm/dd/yyyy"
                ws.cell(current_row, 7).alignment = Alignment(horizontal="center")
            except ValueError:
                ws.cell(current_row, 7).value = created_str[:10]

        ws.cell(current_row, 8).value = info.get("status", "")
        tc = ws.cell(current_row, 9)
        tc.value = total_h if total_h > 0 else 0
        tc.number_format = '0.0"h"'

        if total_h == 0:
            for col_idx in [1, 2, 7, 8, 9]:
                ws.cell(current_row, col_idx).font = Font(color=RED, bold=(col_idx in [1, 9]))

        current_row += 1

    summary_end = current_row - 1
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    ws.cell(current_row, 1).value = "SUMMARY TOTAL"
    ws.cell(current_row, 1).font = Font(bold=True, size=11)
    summary_total = ws.cell(current_row, 9)
    summary_total.value = f"=SUM(I{summary_start}:I{summary_end})"
    summary_total.number_format = '0.0"h"'
    summary_total.font = Font(bold=True, size=11)
    summary_total.fill = blue_fill

    # Column widths
    ws.column_dimensions["A"].width = 15
    for col_letter in "BCDEFGH":
        ws.column_dimensions[col_letter].width = 10
    ws.column_dimensions["I"].width = 12

    out_path = output_dir / f"timesheet_{month}_{year}.xlsx"
    try:
        wb.save(str(out_path))
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        out_path = output_dir / f"timesheet_{month}_{year}_{ts}.xlsx"
        wb.save(str(out_path))
        print(f"  (original file was open in Excel — saved to new file instead)")
    print(f"\nTimesheet written → {out_path}")


def generate_timesheet(extra_args: list[str]) -> None:
    """Fetch worklogs and produce an Excel timesheet for a given month/year."""
    missing = [v for v in ("JIRA_BASE_URL", "JIRA_EMAIL", "JIRA_API_TOKEN") if not os.getenv(v)]
    if missing:
        print("ERROR: Missing required environment variables:", ", ".join(missing))
        sys.exit(1)

    date_range = get_date_range(extra_args)
    print(f"Jira base URL : {JIRA_BASE_URL}")
    print(f"User          : {JIRA_EMAIL}")
    print(
        f"Period        : {date_range['month']:02d}/{date_range['year']}  "
        f"({date_range['start_date']} → {date_range['end_date']})"
    )
    print()

    current_user_profile = get_current_user()
    current_user_id = current_user_profile.get("accountId", JIRA_EMAIL)

    print("Fetching issues with logged hours in range...")
    issues_with_hours = fetch_issues_with_worklogs_in_range(date_range)
    print("Fetching issues assigned during range...")
    issues_assigned = fetch_assigned_issues_in_range(date_range)

    issues_map = {i["key"]: i for i in issues_with_hours}
    for i in issues_assigned:
        issues_map.setdefault(i["key"], i)
    issues = list(issues_map.values())

    print(
        f"Found {len(issues)} issues "
        f"({len(issues_with_hours)} with hours, {len(issues_assigned)} assigned)\n"
    )

    # The inline `worklog` field in search results is capped at 20 entries.
    # Fetch the full worklog history per issue so hours are accurate.
    print("Fetching complete worklogs per issue...", flush=True)
    for idx, issue in enumerate(issues):
        try:
            full_logs = fetch_all_worklogs(issue["key"])
            issue.setdefault("fields", {}).setdefault("worklog", {})["worklogs"] = full_logs
        except Exception:
            pass  # keep whatever inline data was returned
        print(f"  Worklogs: {idx + 1}/{len(issues)}", end="\r", flush=True)
    print()

    total_hours = 0.0
    for issue in issues:
        wl = calculate_worklog_hours_in_range(issue, current_user_id, date_range)
        total_hours += float(wl["total_hours"])
        fields = issue.get("fields", {})
        print(f"{issue['key']} | {(fields.get('issuetype') or {}).get('name', '')}")
        print(f"  {fields.get('summary', '')}")
        print(f"  Status      : {(fields.get('status') or {}).get('name', '')}")
        print(f"  Time Logged : {wl['total_hours']}h ({wl['count']} entries)")
        print()

    print("=" * 80)
    print(f"Total time logged: {total_hours:.2f} hours")
    print("=" * 80)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    generate_excel_report(issues, date_range, current_user_id, OUTPUT_DIR)


def download_other_ticket(issue_key: str) -> None:
    """Download any Jira ticket (regardless of assignment) into the Others/ folder."""
    missing = [v for v in ("JIRA_BASE_URL", "JIRA_EMAIL", "JIRA_API_TOKEN") if not os.getenv(v)]
    if missing:
        print("ERROR: Missing required environment variables:", ", ".join(missing))
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "Others").mkdir(parents=True, exist_ok=True)

    print(f"Jira base URL : {JIRA_BASE_URL}")
    print(f"User          : {JIRA_EMAIL}")
    print(f"Output dir    : {OUTPUT_DIR / 'Others'}")
    print()
    print(f"Fetching ticket {issue_key}...")
    try:
        issue = fetch_single_issue(issue_key)
    except requests.HTTPError as exc:
        print(f"HTTP error fetching {issue_key}: {exc}")
        sys.exit(1)

    print(f"  Processing {issue_key}...")
    try:
        folder, _ = process_issue(issue, OUTPUT_DIR, folder_override="Others")
        print(f"\nTicket {issue_key} saved → {OUTPUT_DIR / folder}/")
    except Exception as exc:
        print(f"ERROR processing {issue_key}: {exc}")
        sys.exit(1)


def _which_with_fallbacks(name: str, fallbacks: list[str]) -> Optional[str]:
    """shutil.which plus explicit probe paths (npm's global dir is often not on PATH)."""
    import shutil
    found = shutil.which(name)
    if found:
        return found
    for candidate in fallbacks:
        p = Path(os.path.expandvars(candidate))
        if p.is_file():
            return str(p)
    return None


def _run_quiet(cmd: list[str], timeout: int = 15) -> tuple[int, str]:
    """Run a command, return (exitcode, first line of output). Never raises."""
    import subprocess
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
        out = (r.stdout or r.stderr or "").strip().splitlines()
        return r.returncode, (out[0] if out else "")
    except FileNotFoundError:
        return 127, "not found"
    except Exception as exc:
        return 1, str(exc)[:80]


def doctor(argv: list[str]) -> None:
    """
    `doctor` — verify the toolchain the ticket lifecycle depends on, before starting.

        python jira_sync.py doctor

    Checks are split into HARD (the lifecycle cannot run) and WARN (a specific step
    will fail later — e.g. no docker means no local E2E, but VB-TEST still works).
    Exit code = number of hard failures, so scripts and the /implement flow can gate
    on it.
    """
    hard: list[str] = []
    warn: list[str] = []

    def ok(label: str, detail: str = "") -> None:
        print(f"  [ OK ] {label}" + (f"  ({detail})" if detail else ""))

    def fail(label: str, fix: str) -> None:
        hard.append(label)
        print(f"  [FAIL] {label}\n         fix: {fix}")

    def note(label: str, fix: str) -> None:
        warn.append(label)
        print(f"  [WARN] {label}\n         {fix}")

    print("jira-sync doctor — ticket-lifecycle toolchain\n")

    # --- .env / Jira ----------------------------------------------------------
    missing = [v for v in ("JIRA_BASE_URL", "JIRA_EMAIL", "JIRA_API_TOKEN") if not os.getenv(v)]
    if missing:
        fail(f".env incomplete: missing {', '.join(missing)}",
             f"copy .env.example to .env beside jira_sync.py and fill it in ({Path(__file__).with_name('.env')})")
    else:
        ok(".env core keys present", JIRA_EMAIL)
        try:
            me = _get("/rest/api/3/myself")
            ok("Jira API reachable, token valid", me.get("displayName", "?"))
        except Exception as exc:
            fail(f"Jira API call failed: {str(exc)[:70]}",
                 "check JIRA_BASE_URL has no trailing path, the token at id.atlassian.com, and the VPN")

    if os.getenv("GITHUB_REPO"):
        ok("GITHUB_REPO set", os.getenv("GITHUB_REPO"))
    else:
        note("GITHUB_REPO not set in .env",
             "PR links in `deliver --pr` and PR-status lookups will be skipped")

    # --- python dependencies (catches "ran with the wrong interpreter") --------
    import importlib.util
    missing_mods = [m for m in ("requests", "dotenv", "openpyxl")
                    if importlib.util.find_spec(m) is None]
    if missing_mods:
        fail(f"python deps missing in THIS interpreter: {', '.join(missing_mods)}",
             "run setup.ps1 beside this script, or `.venv/Scripts/pip install -r requirements.txt`")
    else:
        ok("python deps (requests, dotenv, openpyxl)", sys.executable)

    # --- gh CLI ----------------------------------------------------------------
    gh = _which_with_fallbacks("gh", [r"%ProgramFiles%\GitHub CLI\gh.exe"])
    if not gh:
        fail("gh CLI not installed", "https://cli.github.com — then `gh auth login`")
    else:
        code, _ = _run_quiet([gh, "auth", "status"])
        if code == 0:
            ok("gh CLI authenticated")
            repo = os.getenv("GITHUB_REPO")
            if repo:
                code, push = _run_quiet([gh, "api", f"repos/{repo}", "-q", ".permissions.push"])
                if code == 0 and push == "true":
                    ok(f"push access to {repo}")
                elif code == 0:
                    note(f"no push access to {repo}", "PR creation will fail; ask for write access")
                else:
                    note(f"could not query {repo}", "check the repo slug in .env and your network")
        else:
            fail("gh CLI present but not authenticated", "run `gh auth login`")

    # --- node / newman (global, per the runner convention) ----------------------
    node = _which_with_fallbacks("node", [r"%ProgramFiles%\nodejs\node.exe"])
    if node:
        _, v = _run_quiet([node, "--version"])
        ok("node", v)
    else:
        fail("node not installed", "https://nodejs.org — newman needs it")
    newman = _which_with_fallbacks("newman", [r"%APPDATA%\npm\newman.cmd",
                                              r"%ProgramFiles%\nodejs\newman.cmd"])
    if newman:
        ok("newman (global)", newman)
    else:
        fail("newman not installed globally", "npm install -g newman")

    # --- dotnet / git / docker ---------------------------------------------------
    code, v = _run_quiet(["dotnet", "--version"])
    if code == 0:
        ok("dotnet SDK", v)
    else:
        fail("dotnet SDK not on PATH", "install the .NET Core 3.1 SDK the repo pins in global.json")

    code, v = _run_quiet(["git", "--version"])
    ok("git", v) if code == 0 else fail("git not on PATH", "install Git for Windows")

    code, _ = _run_quiet(["docker", "ps", "-q"], timeout=20)
    if code == 0:
        ok("docker daemon running")
    else:
        note("docker daemon not running",
             "local E2E, seed repair and Redis flush need it; VB-TEST runs do not. Start Docker Desktop + docker\\dev-env-up.bat")

    # --- code repo (when the Claude session exports it) --------------------------
    code_repo = os.getenv("VIVA_CODE_REPO")
    if code_repo and (Path(code_repo) / ".git").is_dir():
        ok("VIVA_CODE_REPO", code_repo)
    elif code_repo:
        note(f"VIVA_CODE_REPO points at '{code_repo}' but there is no .git there",
             "must be the checkout ROOT, not the identically-named source subfolder")
    else:
        note("VIVA_CODE_REPO not set",
             "set it in ~/.claude/settings.json env (see ApiLLM llm/hook-setup.md §3)")

    print(f"\n{len(hard)} hard failure(s), {len(warn)} warning(s).")
    if hard:
        print("Fix the failures above before starting a ticket.")
    sys.exit(len(hard))


# Statuses from which development work may start. Anything else is either finished
# (Done/Closed/UAT...) or explicitly parked (Blocked/PAUSED) and needs a human call.
STARTABLE_STATUSES = {"to do", "backlog", "in progress", "feedback", "to do analyze",
                      "selected for development", "open", "new", "ready for sprint"}

RESEARCH_MARKERS = ("investigation", "spike", "analysis", "research")


def check_ready(argv: list[str]) -> None:
    """
    `ready` — the intake gate for a development ticket.

        python jira_sync.py ready API-9999

    A dev ticket may be started only when it carries the TestCaseReady label AND its
    evidence subtask exists with a non-empty test-case matrix. Everything else is
    reported as a warning for a human decision. Exit code: 0 ready / 1 not ready —
    the /implement flow gates on it. Research tickets are exempt (no matrix needed)
    and reported as such.
    """
    if not argv:
        print("usage: python jira_sync.py ready API-9999")
        sys.exit(2)
    key = argv[0].upper()

    missing = [v for v in ("JIRA_BASE_URL", "JIRA_EMAIL", "JIRA_API_TOKEN") if not os.getenv(v)]
    if missing:
        print("ERROR: Missing required environment variables:", ", ".join(missing))
        print(f"Copy .env.example to .env beside jira_sync.py and fill it in — or run `doctor`.")
        sys.exit(1)

    issue = _get(f"/rest/api/3/issue/{key}",
                 params={"fields": "summary,status,labels,assignee,subtasks,issuetype"})
    f = issue["fields"]
    summary = f.get("summary") or ""
    status = ((f.get("status") or {}).get("name") or "").strip()
    labels = f.get("labels") or []
    assignee = (f.get("assignee") or {}).get("displayName")

    print(f"Ticket   : {key} — {summary}")
    print(f"Status   : {status}")
    print(f"Labels   : {labels}")
    print(f"Assignee : {assignee or 'Unassigned'}\n")

    if any(m in summary.lower() for m in RESEARCH_MARKERS):
        print("Type     : RESEARCH — the TestCaseReady/matrix gate does not apply.")
        print("Follow TICKET-LIFECYCLE.md §4: deliver a written finding + follow-up ticket; no PR, no TestComplete.")
        sys.exit(0)

    blockers: list[str] = []
    warnings: list[str] = []

    if TEST_CASE_READY_LABEL not in labels:
        blockers.append(f"label '{TEST_CASE_READY_LABEL}' is missing — the ticket is NOT ready to start. "
                        "Ask the reporter to define the test cases first.")

    sub_key = None
    for sub in f.get("subtasks") or []:
        if (sub["fields"].get("summary") or "").strip().lower() == EVIDENCE_SUBTASK_SUMMARY.lower():
            sub_key = sub["key"]
            break
    if sub_key:
        sub = _get(f"/rest/api/3/issue/{sub_key}", params={"fields": "description"})
        matrix = field_to_text(sub["fields"].get("description")) or ""
        if len(matrix.strip()) < 40:
            blockers.append(f"evidence subtask {sub_key} exists but its description holds no test-case matrix.")
        else:
            rows = matrix.count("Row ")
            print(f"Matrix   : {sub_key} — {rows or '?'} row(s) found in the subtask description")
    else:
        blockers.append(f"no '{EVIDENCE_SUBTASK_SUMMARY}' subtask — nowhere for the matrix or the evidence.")

    if status.lower() not in STARTABLE_STATUSES:
        warnings.append(f"status '{status}' is not a startable one ({', '.join(sorted(STARTABLE_STATUSES))}). "
                        "If this is a rework return, expect it to be In Progress.")
    if TEST_COMPLETE_LABEL in labels:
        warnings.append(f"'{TEST_COMPLETE_LABEL}' already present — this ticket was delivered before. "
                        "Treat as rework: re-run the affected cases and re-deliver.")
    if assignee and JIRA_EMAIL and JIRA_EMAIL.split("@")[0].lower() not in assignee.lower().replace(" ", "."):
        warnings.append(f"assigned to '{assignee}' — confirm it is yours to take.")

    for b in blockers:
        print(f"  [BLOCK] {b}")
    for w in warnings:
        print(f"  [WARN ] {w}")

    if blockers:
        print(f"\nNOT READY — {len(blockers)} blocker(s). Do not start; report them to the requester.")
        sys.exit(1)
    print(f"\nREADY — proceed with ANALYZE-TASK.md. ({len(warnings)} warning(s) above need a human answer.)")
    sys.exit(0)


def deliver_ticket(argv: list[str]) -> None:
    """
    `deliver` — the write half of the dev-ticket lifecycle.

        python jira_sync.py deliver API-9999 \
            --evidence docs/evidence/API-9999 \
            --pr 2395 \
            --comment "TC01 and TC02 executed in VB-TEST, evidence attached." \
            [--status "In review"] [--no-label] [--keep-assignee] [--dry-run]

    Order is deliberate: attach evidence FIRST, then comment, then label, then
    transition, then unassign. If a step fails the ticket is left in the most
    truthful state reachable — evidence present without the label is recoverable;
    the label claiming completion without evidence is a lie.

    TestCaseReady is NOT removed: every completed ticket in this project carries
    both labels, so TestComplete is added alongside it.

    Unassigning is part of the standard handoff — an In-review ticket left assigned
    looks owned, so QA does not grab it. This applies to rework re-deliveries too;
    --keep-assignee opts out only when a direct handback to a specific person was
    agreed.
    """
    import argparse
    parser = argparse.ArgumentParser(prog="jira_sync.py deliver")
    parser.add_argument("issue_key")
    parser.add_argument("--evidence", metavar="PATH",
                        help="File, or directory whose files are all uploaded, to the evidence subtask.")
    parser.add_argument("--pr", type=int, help="PR number, used in the comment.")
    parser.add_argument("--repo", default=GITHUB_REPO, help="owner/name for the PR link.")
    parser.add_argument("--comment", help="Extra text appended to the delivery comment.")
    parser.add_argument("--status", default=IN_REVIEW_STATUS, help=f"Target status (default: {IN_REVIEW_STATUS}).")
    parser.add_argument("--no-label", action="store_true", help=f"Do not add {TEST_COMPLETE_LABEL}.")
    parser.add_argument("--no-transition", action="store_true", help="Do not change the status.")
    parser.add_argument("--keep-assignee", action="store_true",
                        help="Skip the unassign step (rework loop: the ticket stays yours).")
    parser.add_argument("--dry-run", action="store_true", help="Report what would happen, change nothing.")
    args = parser.parse_args(argv)

    key = args.issue_key.upper()
    missing = [v for v in ("JIRA_BASE_URL", "JIRA_EMAIL", "JIRA_API_TOKEN") if not os.getenv(v)]
    if missing:
        print("ERROR: Missing required environment variables:", ", ".join(missing))
        sys.exit(1)

    issue = _get(f"/rest/api/3/issue/{key}", params={"fields": "summary,status,labels,assignee"})
    f = issue["fields"]
    print(f"Ticket   : {key} — {f['summary']}")
    print(f"Status   : {(f.get('status') or {}).get('name')}")
    print(f"Labels   : {f.get('labels')}")
    if args.dry_run:
        print("\n-- DRY RUN, nothing will be modified --")

    # 1. Evidence -------------------------------------------------------------
    if args.evidence:
        src = Path(args.evidence)
        files = sorted(p for p in (src.iterdir() if src.is_dir() else [src]) if p.is_file())
        if not files:
            print(f"  ⚠  No files found at {src}")
        else:
            sub = jira_find_or_create_evidence_subtask(key) if not args.dry_run else "(dry-run)"
            print(f"  Evidence → {sub}  ({len(files)} file(s))")
            for p in files:
                print(f"    {'would upload' if args.dry_run else '→'} {p.name}")
            if not args.dry_run:
                if not sub:
                    print("  ⛔ Could not resolve or create the evidence subtask — aborting before "
                          "comment/label/transition/unassign. Fix and re-run (deliver is idempotent).")
                    sys.exit(1)
                uploaded = jira_upload_attachments(sub, files)
                if len(uploaded) < len(files):
                    print(f"  ⛔ {len(files) - len(uploaded)} of {len(files)} upload(s) failed — aborting "
                          "before comment/label/transition/unassign: a label claiming completion with no "
                          "evidence behind it is a lie. Fix and re-run (deliver is idempotent).")
                    sys.exit(1)

    # 2. Comment --------------------------------------------------------------
    lines = []
    if args.pr:
        repo = args.repo or ""
        lines.append(f"PR: https://github.com/{repo}/pull/{args.pr}" if repo else f"PR: #{args.pr}")
    if args.comment:
        lines.append(args.comment)
    if lines:
        text = "\n\n".join(lines)
        print(f"  Comment  : {text.splitlines()[0][:70]}…" if len(text) > 70 else f"  Comment  : {text}")
        if not args.dry_run:
            jira_add_comment(key, text)

    # 3. Label ----------------------------------------------------------------
    if not args.no_label:
        current = f.get("labels") or []
        if TEST_COMPLETE_LABEL in current:
            print(f"  Label    : {TEST_COMPLETE_LABEL} already present")
        else:
            print(f"  Label    : + {TEST_COMPLETE_LABEL}  (keeping {TEST_CASE_READY_LABEL})")
            if not args.dry_run:
                print(f"             now: {jira_update_labels(key, add=[TEST_COMPLETE_LABEL])}")

    # 4. Transition -----------------------------------------------------------
    if not args.no_transition:
        now = ((f.get("status") or {}).get("name") or "").strip().lower()
        if now == args.status.strip().lower():
            print(f"  Status   : already '{args.status}'")
        else:
            print(f"  Status   : → {args.status}")
            if not args.dry_run and not jira_transition(key, args.status):
                print("  ⛔ Transition failed — leaving the assignee untouched so the ticket state "
                      "stays truthful. Fix and re-run (deliver is idempotent).")
                sys.exit(1)

    # 5. Unassign (QA handoff) --------------------------------------------------
    if args.keep_assignee:
        print("  Assignee : kept (rework loop)")
    else:
        assignee_field = _get(f"/rest/api/3/issue/{key}", params={"fields": "assignee"}) \
            .get("fields", {}).get("assignee") if not args.dry_run else f.get("assignee")
        if assignee_field is None:
            print("  Assignee : already unassigned")
        else:
            print(f"  Assignee : {assignee_field.get('displayName')} → Unassigned (QA can grab it)")
            if not args.dry_run:
                jira_unassign(key)

    print("\nDone." if not args.dry_run else "\nDry run complete — nothing changed.")


if __name__ == "__main__":
    _KNOWN = {"timesheet", "ticket", "deliver", "doctor", "ready"}
    if len(sys.argv) > 1 and sys.argv[1] in _KNOWN:
        _cmd = sys.argv[1]
        if _cmd == "timesheet":
            generate_timesheet(sys.argv[2:])
        elif _cmd == "doctor":
            doctor(sys.argv[2:])
        elif len(sys.argv) > 2:
            if _cmd == "ticket":
                download_other_ticket(sys.argv[2].upper())
            elif _cmd == "deliver":
                deliver_ticket(sys.argv[2:])
            else:
                check_ready(sys.argv[2:])
        else:
            # A subcommand missing its issue key must NOT fall through to the full
            # sync (which can write to Jira) — that silent fallback was a footgun.
            print(f"ERROR: '{_cmd}' requires an issue key, e.g.  python jira_sync.py {_cmd} API-9999")
            sys.exit(2)
    elif len(sys.argv) > 1 and not sys.argv[1].startswith("-"):
        print(f"ERROR: unknown subcommand '{sys.argv[1]}'. "
              f"Known: {', '.join(sorted(_KNOWN))} — or flags (--from/--to/--all) for the sync.")
        sys.exit(2)
    else:
        main()
